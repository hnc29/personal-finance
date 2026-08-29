"""Service for generating bank-statement formatted export datasets, XLSX, and CSV."""
from __future__ import annotations

import csv
import datetime
import io
import re
from typing import Any

from openpyxl import Workbook  # type: ignore[import-untyped]
from openpyxl.styles import (  # type: ignore[import-untyped]
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from sqlalchemy import func, select
from sqlalchemy.orm import Session

from app.core.money import scaled_to_money
from app.models.account import Account, AccountType
from app.models.category import Category
from app.models.ledger import AccountEntry, FinancialEvent, FinancialEventType

_EVENT_TYPE_VI_LABELS: dict[FinancialEventType, str] = {
    FinancialEventType.EXPENSE: "Chi tiêu",
    FinancialEventType.INCOME: "Thu nhập",
    FinancialEventType.TRANSFER: "Transfer",
    FinancialEventType.CREDIT_CARD_PAYMENT: "Thanh toán thẻ",
    FinancialEventType.INTEREST: "Hạch toán lãi",
    FinancialEventType.SAVINGS_DEPOSIT: "Nộp tiền tiết kiệm",
    FinancialEventType.SAVINGS_WITHDRAWAL: "Rút tiền tiết kiệm",
    FinancialEventType.ASSET_PURCHASE: "Mua tài sản",
    FinancialEventType.ASSET_SALE: "Bán tài sản",
    FinancialEventType.ADJUSTMENT: "Điều chỉnh",
}

_SPREADSHEET_FORMULA_PREFIXES = ("=", "+", "-", "@", "\t", "\r", "\n")


def _sanitize_statement_text(value: str) -> str:
    """Neutralize untrusted text that spreadsheets may interpret as a formula."""
    if value.startswith(_SPREADSHEET_FORMULA_PREFIXES):
        return f"'{value}"
    return value


def _extract_ref_number(event: FinancialEvent) -> str:
    """Extract or synthesize a bank reference number (e.g. FT..., AZ-...)."""
    if event.note:
        match = re.search(r"\b(FT\d+[\w\\/]*|AZ-\d+)\b", event.note)
        if match:
            return match.group(1)
    return f"TX{event.id:08d}"


def get_statement_data(
    db: Session,
    account_id: int | None,
    start_date: datetime.date | None,
    end_date: datetime.date | None,
) -> dict[str, Any]:
    """Calculate opening balance, running balances, and statement rows for a date range."""
    # 1. Target Account Info
    account_obj = db.get(Account, account_id) if account_id is not None else None
    account_name = account_obj.name if account_obj else "Tất cả tài khoản"
    account_type_label = (
        "Tài khoản thanh toán"
        if account_obj and account_obj.account_type in (AccountType.BANK, AccountType.EWALLET)
        else ("Thẻ tín dụng" if account_obj and account_obj.account_type == AccountType.CREDIT_CARD else "Sổ tổng hợp")
    )
    currency = account_obj.currency if account_obj else "VND"

    # 2. Calculate Opening Balance (sum of all entries before start_date)
    opening_scaled = 0
    if start_date is not None:
        open_stmt = (
            select(func.coalesce(func.sum(AccountEntry.amount_scaled), 0))
            .join(FinancialEvent, AccountEntry.financial_event_id == FinancialEvent.id)
            .where(FinancialEvent.transaction_date < start_date)
        )
        if account_id is not None:
            open_stmt = open_stmt.where(AccountEntry.account_id == account_id)
        opening_scaled = int(db.scalar(open_stmt) or 0)

    # 3. Query transactions in range and category map
    category_map = {c.id: c.name for c in db.scalars(select(Category)).all()}

    tx_stmt = (
        select(AccountEntry, FinancialEvent)
        .join(FinancialEvent, AccountEntry.financial_event_id == FinancialEvent.id)
        .order_by(FinancialEvent.transaction_date.asc(), FinancialEvent.id.asc(), AccountEntry.id.asc())
    )
    if account_id is not None:
        tx_stmt = tx_stmt.where(AccountEntry.account_id == account_id)
    if start_date is not None:
        tx_stmt = tx_stmt.where(FinancialEvent.transaction_date >= start_date)
    if end_date is not None:
        tx_stmt = tx_stmt.where(FinancialEvent.transaction_date <= end_date)

    raw_pairs = db.execute(tx_stmt).all()

    # 4. Process rows and compute running balance
    current_running_scaled = opening_scaled
    total_in_scaled = 0
    total_out_scaled = 0
    rows: list[dict[str, Any]] = []

    for entry, event in raw_pairs:
        amt_scaled = entry.amount_scaled
        current_running_scaled += amt_scaled
        if amt_scaled > 0:
            total_in_scaled += amt_scaled
        else:
            total_out_scaled += abs(amt_scaled)

        amt_decimal = scaled_to_money(amt_scaled)
        running_decimal = scaled_to_money(current_running_scaled)
        ref_no = _extract_ref_number(event)
        event_label = _EVENT_TYPE_VI_LABELS.get(event.event_type, event.event_type.value)

        # Build clean description
        description_parts = []
        if event.note:
            description_parts.append(event.note)
        elif event.category_id and event.category_id in category_map:
            description_parts.append(category_map[event.category_id])
        if event.payee_text:
            description_parts.append(f"({event.payee_text})")
        description = " ".join(description_parts) or event_label

        rows.append({
            "id": event.id,
            "entry_id": entry.id,
            "account_id": entry.account_id,
            "transaction_date": event.transaction_date.isoformat(),
            "effective_date": event.transaction_date.isoformat(),
            "event_type": event.event_type.value,
            "event_type_label": event_label,
            "description": description,
            "ref_no": ref_no,
            "amount": str(amt_decimal),
            "amount_scaled": amt_scaled,
            "running_balance": str(running_decimal),
            "running_balance_scaled": current_running_scaled,
        })

    opening_balance_decimal = scaled_to_money(opening_scaled)
    closing_balance_decimal = scaled_to_money(current_running_scaled)
    total_in_decimal = scaled_to_money(total_in_scaled)
    total_out_decimal = scaled_to_money(total_out_scaled)

    return {
        "account": {
            "id": account_obj.id if account_obj else None,
            "name": account_name,
            "account_type": account_obj.account_type.value if account_obj else "ALL",
            "account_type_label": account_type_label,
            "currency": currency,
        },
        "period": {
            "start_date": start_date.isoformat() if start_date else None,
            "end_date": end_date.isoformat() if end_date else None,
        },
        "opening_balance": str(opening_balance_decimal),
        "closing_balance": str(closing_balance_decimal),
        "total_in": str(total_in_decimal),
        "total_out": str(total_out_decimal),
        "transaction_count": len(rows),
        "transactions": rows,
    }


def generate_statement_xlsx(
    db: Session,
    account_id: int | None,
    start_date: datetime.date | None,
    end_date: datetime.date | None,
) -> io.BytesIO:
    """Generate a styled Excel workbook formatted as a formal bank statement."""
    data = get_statement_data(db, account_id, start_date, end_date)
    wb = Workbook()
    ws = wb.active
    ws.title = "Sao Kê Tài Khoản"

    # Styling definitions
    title_font = Font(name="Arial", size=15, bold=True, color="003366")
    meta_bold_font = Font(name="Arial", size=10, bold=True)
    meta_font = Font(name="Arial", size=10)
    header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    row_font = Font(name="Arial", size=9)
    green_font = Font(name="Arial", size=9, color="16A34A", bold=True)
    red_font = Font(name="Arial", size=9, color="DC2626", bold=True)

    thin_border = Border(
        left=Side(style="thin", color="E2E8F0"),
        right=Side(style="thin", color="E2E8F0"),
        top=Side(style="thin", color="E2E8F0"),
        bottom=Side(style="thin", color="E2E8F0"),
    )

    # 1. Header Block
    ws["A1"] = "SAO KÊ TÀI KHOẢN"
    ws["A1"].font = title_font
    ws.merge_cells("A1:G1")

    p_start = data["period"]["start_date"] or "Từ đầu"
    p_end = data["period"]["end_date"] or "Đến nay"
    ws["A2"] = f"Thời gian truy vấn: {p_start} - {p_end}"
    ws["A2"].font = meta_font
    ws.merge_cells("A2:G2")

    # 2. Account Metadata Block
    ws["A4"] = "Tài khoản:"
    ws["A4"].font = meta_bold_font
    ws["B4"] = _sanitize_statement_text(data["account"]["name"])
    ws["B4"].font = meta_font

    ws["D4"] = "Loại tài khoản:"
    ws["D4"].font = meta_bold_font
    ws["E4"] = data["account"]["account_type_label"]
    ws["E4"].font = meta_font

    ws["A5"] = "Số dư đầu kỳ:"
    ws["A5"].font = meta_bold_font
    ws["B5"] = f"{float(data['opening_balance']):,.0f} {data['account']['currency']}"
    ws["B5"].font = meta_font

    ws["D5"] = "Số dư cuối kỳ:"
    ws["D5"].font = meta_bold_font
    ws["E5"] = f"{float(data['closing_balance']):,.0f} {data['account']['currency']}"
    ws["E5"].font = meta_font

    ws["A6"] = "Tổng tiền vào (+):"
    ws["A6"].font = meta_bold_font
    ws["B6"] = f"+{float(data['total_in']):,.0f} {data['account']['currency']}"
    ws["B6"].font = Font(name="Arial", size=10, color="16A34A", bold=True)

    ws["D6"] = "Tổng tiền ra (-):"
    ws["D6"].font = meta_bold_font
    ws["E6"] = f"-{float(data['total_out']):,.0f} {data['account']['currency']}"
    ws["E6"].font = Font(name="Arial", size=10, color="DC2626", bold=True)

    # 3. Table Headers
    headers = ["Ngày", "Ngày hiệu lực", "Loại giao dịch", "Nội dung", "Ref#", "Số tiền giao dịch", "Số dư"]
    ws.append([])  # Row 7 empty
    ws.append(headers)  # Row 8
    header_row_idx = 8

    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=header_row_idx, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center" if col_idx in (1, 2, 3, 5) else ("right" if col_idx in (6, 7) else "left"), vertical="center")

    # 4. Table Rows
    curr_row = header_row_idx + 1
    for tx in data["transactions"]:
        amt = float(tx["amount"])
        amt_str = f"+{amt:,.0f}" if amt > 0 else f"{amt:,.0f}"
        run_bal_str = f"{float(tx['running_balance']):,.0f}"

        row_values = [
            tx["transaction_date"],
            tx["effective_date"],
            tx["event_type_label"],
            _sanitize_statement_text(tx["description"]),
            _sanitize_statement_text(tx["ref_no"]),
            amt_str,
            run_bal_str,
        ]
        ws.append(row_values)

        ws.cell(row=curr_row, column=1).alignment = Alignment(horizontal="center")
        ws.cell(row=curr_row, column=2).alignment = Alignment(horizontal="center")
        ws.cell(row=curr_row, column=3).alignment = Alignment(horizontal="center")
        ws.cell(row=curr_row, column=4).alignment = Alignment(horizontal="left", wrap_text=True)
        ws.cell(row=curr_row, column=5).alignment = Alignment(horizontal="center")
        ws.cell(row=curr_row, column=6).alignment = Alignment(horizontal="right")
        ws.cell(row=curr_row, column=7).alignment = Alignment(horizontal="right")

        for c in range(1, 8):
            cell = ws.cell(row=curr_row, column=c)
            cell.font = green_font if c == 6 and amt > 0 else (red_font if c == 6 and amt < 0 else row_font)
            cell.border = thin_border

        curr_row += 1

    # Set Column Widths
    col_widths = {"A": 13, "B": 13, "C": 16, "D": 45, "E": 24, "F": 20, "G": 20}
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out


def generate_statement_csv(
    db: Session,
    account_id: int | None,
    start_date: datetime.date | None,
    end_date: datetime.date | None,
) -> io.StringIO:
    """Generate a CSV string formatted with statement headers and running balances in UTF-8 BOM."""
    data = get_statement_data(db, account_id, start_date, end_date)
    out = io.StringIO()
    writer = csv.writer(out)

    # Header metadata
    writer.writerow(
        ["SAO KÊ TÀI KHOẢN", _sanitize_statement_text(data["account"]["name"])]
    )
    writer.writerow(["Thời gian truy vấn", f"{data['period']['start_date'] or 'Từ đầu'} - {data['period']['end_date'] or 'Đến nay'}"])
    currency = _sanitize_statement_text(data["account"]["currency"])
    writer.writerow(["Số dư đầu kỳ", data["opening_balance"], currency])
    writer.writerow(["Số dư cuối kỳ", data["closing_balance"], currency])
    writer.writerow([])

    # Table headers
    writer.writerow(["Ngày", "Ngày hiệu lực", "Loại giao dịch", "Nội dung", "Ref#", "Số tiền giao dịch", "Số dư"])
    for tx in data["transactions"]:
        amt = float(tx["amount"])
        amt_str = f"+{amt:,.0f}" if amt > 0 else f"{amt:,.0f}"
        run_bal_str = f"{float(tx['running_balance']):,.0f}"
        writer.writerow([
            tx["transaction_date"],
            tx["effective_date"],
            tx["event_type_label"],
            _sanitize_statement_text(tx["description"]),
            _sanitize_statement_text(tx["ref_no"]),
            amt_str,
            run_bal_str,
        ])
    return out
