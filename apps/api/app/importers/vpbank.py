"""VPBank XLSX statement adapter."""

from __future__ import annotations

import io
import re
import unicodedata
from collections.abc import Iterable
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from typing import BinaryIO

from openpyxl import load_workbook  # type: ignore[import-untyped]

from app.core.money import money_to_scaled, scaled_to_money
from app.importers.bank_statement import NormalizedBankStatementRow

_FIELD_ALIASES = {
    "transaction_date": ("ngay giao dich", "transaction date"),
    "effective_date": (
        "ngay hieu luc",
        "ngay gia tri",
        "ngay hach toan",
        "effective date",
        "value date",
        "posting date",
    ),
    "reference": ("so tham chieu", "so but toan", "reference", "reference number"),
    "description": (
        "dien giai",
        "mo ta",
        "noi dung",
        "description",
        "transaction details",
    ),
    "debit": ("ghi no", "so tien ghi no", "debit", "debit amount"),
    "credit": ("ghi co", "so tien ghi co", "credit", "credit amount"),
    "running_balance": ("so du", "so du cuoi", "balance", "running balance"),
}
_REQUIRED_FIELDS = frozenset(("transaction_date", "debit", "credit", "running_balance"))
_DATE_FORMATS = ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d", "%d/%m/%Y %H:%M:%S")


class VPBankStatementAdapter:
    """Parse VPBank workbooks into normalized statement rows."""

    def parse(self, source: bytes | BinaryIO) -> tuple[NormalizedBankStatementRow, ...]:
        data = source if isinstance(source, bytes) else source.read()
        if not isinstance(data, bytes):
            raise TypeError("source must provide XLSX bytes")

        buffer = io.BytesIO(data)
        workbook = load_workbook(buffer, read_only=True, data_only=True)
        try:
            sheet = workbook.active
            header_row, columns = _find_header(sheet.iter_rows(values_only=True))
            rows: list[NormalizedBankStatementRow] = []
            for row_number, values in enumerate(
                sheet.iter_rows(min_row=header_row + 1, values_only=True),
                start=header_row + 1,
            ):
                if not any(
                    value is not None and str(value).strip() for value in values
                ):
                    continue
                transaction_value = _cell(values, columns["transaction_date"])
                if transaction_value is None or str(transaction_value).strip() == "":
                    continue

                debit = _money(_cell(values, columns["debit"]), row_number, "debit")
                credit = _money(_cell(values, columns["credit"]), row_number, "credit")
                _validate_transaction_amount(debit, row_number, "debit")
                _validate_transaction_amount(credit, row_number, "credit")
                if debit is not None and credit is not None:
                    raise ValueError(
                        f"both debit and credit present at source row {row_number}"
                    )
                if debit is None and credit is None:
                    raise ValueError(
                        f"missing debit and credit at source row {row_number}"
                    )

                if credit is not None:
                    signed_amount = credit
                else:
                    assert debit is not None
                    signed_amount = -debit
                rows.append(
                    NormalizedBankStatementRow(
                        source_row_number=row_number,
                        transaction_date=_statement_date(
                            transaction_value, row_number, "transaction date"
                        ),
                        effective_date=_optional_date(
                            _field(values, columns, "effective_date"),
                            row_number,
                            "effective date",
                        ),
                        reference=_text(_field(values, columns, "reference")),
                        description=_text(_field(values, columns, "description")),
                        debit=debit,
                        credit=credit,
                        signed_amount=signed_amount,
                        running_balance=_money(
                            _cell(values, columns["running_balance"]),
                            row_number,
                            "running balance",
                        ),
                    )
                )
            return tuple(rows)
        finally:
            workbook.close()
            buffer.close()


def _find_header(rows: Iterable[tuple[object, ...]]) -> tuple[int, dict[str, int]]:
    for row_number, values in enumerate(rows, start=1):
        columns: dict[str, int] = {}
        for index, value in enumerate(values):
            normalized = _normalized_header(value)
            for field, aliases in _FIELD_ALIASES.items():
                if normalized in aliases:
                    columns[field] = index
                    break
        if _REQUIRED_FIELDS <= columns.keys():
            return row_number, columns
        if row_number >= 50:
            break
    raise ValueError("VPBank statement headers not found")


def _normalized_header(value: object) -> str:
    text = unicodedata.normalize("NFD", "" if value is None else str(value))
    text = "".join(
        character for character in text if not unicodedata.combining(character)
    )
    return re.sub(r"\s+", " ", text.replace("đ", "d").replace("Đ", "D")).strip().lower()


def _cell(values: tuple[object, ...], index: int) -> object | None:
    return values[index] if index < len(values) else None


def _field(
    values: tuple[object, ...], columns: dict[str, int], field: str
) -> object | None:
    index = columns.get(field)
    return None if index is None else _cell(values, index)


def _text(value: object | None) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _statement_date(value: object, row_number: int, field: str) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    for date_format in _DATE_FORMATS:
        try:
            # Statement dates are intentionally calendar dates without a timezone.
            return datetime.strptime(text, date_format).date()  # noqa: DTZ007
        except ValueError:
            pass
    raise ValueError(f"invalid {field} at source row {row_number}")


def _optional_date(value: object | None, row_number: int, field: str) -> date | None:
    if value is None or str(value).strip() == "":
        return None
    return _statement_date(value, row_number, field)


def _money(value: object | None, row_number: int, field: str) -> Decimal | None:
    if value is None or str(value).strip() in ("", "-"):
        return None
    if isinstance(value, float):
        raise ValueError(  # noqa: TRY004 - normalize source-cell failures uniformly.
            f"invalid {field} at source row {row_number}"
        )
    text = str(value).strip().replace(" ", "").replace(",", "")
    try:
        amount = Decimal(text)
        return scaled_to_money(money_to_scaled(amount))
    except (InvalidOperation, ValueError) as exc:
        raise ValueError(f"invalid {field} at source row {row_number}") from exc


def _validate_transaction_amount(
    value: Decimal | None, row_number: int, field: str
) -> None:
    if value is not None and value < 0:
        raise ValueError(f"negative {field} at source row {row_number}")
