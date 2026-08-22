"""Money Lover XLSX source adapter."""

from __future__ import annotations

import hashlib
import io
import json
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal
from typing import BinaryIO

from openpyxl import load_workbook  # type: ignore[import-untyped]

from app.core.money import money_to_scaled

WORKSHEET = "Sổ giao dịch"
HEADERS = (
    "Id",
    "Ngày",
    "Nhóm",
    "Số tiền",
    "Đơn vị tiền tệ",
    "Ví",
    "Ghi chú",
    "Với",
    "Sự kiện",
    "Không tính vào báo cáo",
    "Thành viên",
)


@dataclass(frozen=True)
class ParsedMoneyLoverRow:
    source_row_number: int
    source_row_id: str | None
    raw_payload: dict[str, object]
    transaction_date: date | None
    amount: Decimal | None
    amount_scaled: int | None


@dataclass(frozen=True)
class ParsedMoneyLoverFile:
    file_sha256: str
    rows: tuple[ParsedMoneyLoverRow, ...]


def _json_value(value: object) -> object:
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, Decimal):
        return str(value)
    return value


def parse_moneylover_xlsx(source: bytes | BinaryIO) -> ParsedMoneyLoverFile:
    data = source if isinstance(source, bytes) else source.read()
    if not isinstance(data, bytes):
        raise TypeError("source must provide XLSX bytes")
    digest = hashlib.sha256(data).hexdigest()
    workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    try:
        if WORKSHEET not in workbook.sheetnames:
            raise ValueError(f"missing required worksheet: {WORKSHEET}")
        sheet = workbook[WORKSHEET]
        header_values = tuple(cell.value for cell in next(sheet.iter_rows()))
        if len(header_values) < len(HEADERS) or tuple(header_values[: len(HEADERS)]) != HEADERS:
            raise ValueError("invalid Sổ giao dịch headers")
        source_headers = tuple(str(value) for value in header_values)
        rows: list[ParsedMoneyLoverRow] = []
        for row_number, cells in enumerate(sheet.iter_rows(min_row=2), start=2):
            values = [_json_value(cell.value) for cell in cells]
            payload = dict(zip(source_headers, values, strict=False))
            raw_amount = cells[3].value if len(cells) > 3 else None
            amount = None if raw_amount is None else Decimal(str(raw_amount))
            transaction_date = cells[1].value if len(cells) > 1 else None
            if isinstance(transaction_date, datetime):
                transaction_date = transaction_date.date()
            if transaction_date is not None and not isinstance(transaction_date, date):
                raise ValueError(f"invalid date at source row {row_number}")
            source_id = cells[0].value if cells else None
            rows.append(ParsedMoneyLoverRow(
                row_number,
                None if source_id is None else str(source_id),
                payload,
                transaction_date,
                amount,
                None if amount is None else money_to_scaled(amount),
            ))
        return ParsedMoneyLoverFile(digest, tuple(rows))
    finally:
        workbook.close()


def raw_payload_text(payload: dict[str, object]) -> str:
    return json.dumps(payload, ensure_ascii=False, sort_keys=True, separators=(",", ":"))


def semantic_fingerprint(payload: dict[str, object]) -> str:
    """Return a stable fingerprint for a row's source semantics.

    This is informational only: callers must not use it to merge or discard rows.
    """
    return hashlib.sha256(raw_payload_text(payload).encode("utf-8")).hexdigest()
