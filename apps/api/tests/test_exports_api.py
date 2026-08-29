"""API tests for the /exports endpoints' account/date filters.

Regression test for the user report (2026-08-26): "Chức năng xuất dữ liệu:
Lựa chọn tài khoản, ngày bắt đầu, ngày kết thúc" -- exports used to have no
way to scope which accounts/dates were included. Runs Alembic against a
disposable temp-file SQLite database (never data/finance.db), matching the
pattern in tests/test_assets_api.py. All data is synthetic.
"""

import csv
import io
import os
import subprocess
from collections.abc import Iterator
from decimal import Decimal
from pathlib import Path

import pytest
from fastapi.testclient import TestClient
from sqlalchemy import create_engine, text
from sqlalchemy.orm import Session, sessionmaker

from app.core.database import get_db
from app.main import app
from app.services.statement_export import _sanitize_statement_text


@pytest.fixture
def client(tmp_path: Path) -> Iterator[TestClient]:
    database_path = tmp_path / "synthetic-exports.db"
    env = os.environ.copy()
    env["PF_DATABASE_PATH"] = str(database_path)
    subprocess.run(
        ["alembic", "upgrade", "head"], check=True, env=env, capture_output=True
    )
    engine = create_engine(
        f"sqlite:///{database_path}", connect_args={"check_same_thread": False}
    )
    session_factory = sessionmaker(bind=engine, autoflush=False, expire_on_commit=False)

    def override_get_db() -> Iterator[Session]:
        db = session_factory()
        try:
            db.execute(text("PRAGMA foreign_keys=ON"))
            yield db
        finally:
            db.close()

    app.dependency_overrides[get_db] = override_get_db
    try:
        yield TestClient(app)
    finally:
        app.dependency_overrides.clear()
        engine.dispose()


def _make_account(client: TestClient, name: str) -> int:
    response = client.post(
        "/api/v1/accounts",
        json={"name": name, "account_type": "CASH", "currency": "VND"},
    )
    assert response.status_code == 201
    return response.json()["id"]


def _make_expense(client: TestClient, account_id: int, date: str, amount: str) -> None:
    response = client.post(
        "/api/v1/financial-events",
        json={
            "event_type": "EXPENSE",
            "transaction_date": date,
            "entries": [{"account_id": account_id, "amount": amount}],
        },
    )
    assert response.status_code == 201


def _csv_data_rows(body: bytes) -> list[list[str]]:
    rows = list(csv.reader(io.StringIO(body.decode())))
    return rows[1:]  # drop header


def test_export_csv_unfiltered_includes_every_entry(client: TestClient) -> None:
    a = _make_account(client, "A")
    b = _make_account(client, "B")
    _make_expense(client, a, "2026-01-10", "-10.0000")
    _make_expense(client, b, "2026-02-10", "-20.0000")

    response = client.get("/api/v1/exports/events.csv")
    assert response.status_code == 200
    assert len(_csv_data_rows(response.content)) == 2


def test_export_csv_filters_by_account(client: TestClient) -> None:
    a = _make_account(client, "A")
    b = _make_account(client, "B")
    _make_expense(client, a, "2026-01-10", "-10.0000")
    _make_expense(client, b, "2026-01-10", "-20.0000")

    response = client.get(f"/api/v1/exports/events.csv?account_id={a}")
    assert response.status_code == 200
    rows = _csv_data_rows(response.content)
    assert len(rows) == 1
    assert rows[0][3] == str(a)


def test_export_csv_filters_by_date_range(client: TestClient) -> None:
    a = _make_account(client, "A")
    _make_expense(client, a, "2026-01-05", "-10.0000")
    _make_expense(client, a, "2026-03-15", "-20.0000")
    _make_expense(client, a, "2026-06-01", "-30.0000")

    response = client.get(
        "/api/v1/exports/events.csv?start_date=2026-02-01&end_date=2026-04-01"
    )
    assert response.status_code == 200
    rows = _csv_data_rows(response.content)
    assert len(rows) == 1
    assert rows[0][1] == "2026-03-15"


def test_export_csv_rejects_start_after_end(client: TestClient) -> None:
    response = client.get(
        "/api/v1/exports/events.csv?start_date=2026-06-01&end_date=2026-01-01"
    )
    assert response.status_code == 400


def test_export_xlsx_filters_by_account(client: TestClient) -> None:
    from openpyxl import load_workbook

    a = _make_account(client, "A")
    b = _make_account(client, "B")
    _make_expense(client, a, "2026-01-10", "-10.0000")
    _make_expense(client, b, "2026-01-10", "-20.0000")

    response = client.get(f"/api/v1/exports/events.xlsx?account_id={b}")
    assert response.status_code == 200
    wb = load_workbook(io.BytesIO(response.content))
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    assert len(rows) == 1
    assert rows[0][3] == b


def test_export_statement_data_and_running_balance(client: TestClient) -> None:
    a = _make_account(client, "VPBank")
    # Opening balance before 2026-07-01: +1,000,000
    client.post(
        "/api/v1/financial-events",
        json={
            "event_type": "INCOME",
            "transaction_date": "2026-06-15",
            "entries": [{"account_id": a, "amount": "1000000.0000"}],
        },
    )
    # Transactions in July
    _make_expense(client, a, "2026-07-10", "-200000.0000")
    client.post(
        "/api/v1/financial-events",
        json={
            "event_type": "INTEREST",
            "transaction_date": "2026-07-20",
            "entries": [{"account_id": a, "amount": "50000.0000"}],
        },
    )

    response = client.get(
        f"/api/v1/exports/statement/data?account_id={a}&start_date=2026-07-01&end_date=2026-07-31"
    )
    assert response.status_code == 200
    data = response.json()
    assert data["account"]["name"] == "VPBank"
    assert Decimal(data["opening_balance"]) == Decimal("1000000.0000")
    assert Decimal(data["closing_balance"]) == Decimal("850000.0000")
    assert Decimal(data["total_in"]) == Decimal("50000.0000")
    assert Decimal(data["total_out"]) == Decimal("200000.0000")
    assert len(data["transactions"]) == 2
    # Row 1: -200,000 -> running = 800,000
    assert data["transactions"][0]["amount"] == "-200000.0000"
    assert data["transactions"][0]["running_balance"] == "800000.0000"
    # Row 2: +50,000 -> running = 850,000
    assert data["transactions"][1]["amount"] == "50000.0000"
    assert data["transactions"][1]["running_balance"] == "850000.0000"


def test_export_statement_xlsx_and_csv(client: TestClient) -> None:
    from openpyxl import load_workbook

    a = _make_account(client, "VPBank")
    _make_expense(client, a, "2026-08-01", "-50000.0000")

    # Test XLSX
    resp_xlsx = client.get(f"/api/v1/exports/statement.xlsx?account_id={a}")
    assert resp_xlsx.status_code == 200
    wb = load_workbook(io.BytesIO(resp_xlsx.content))
    ws = wb.active
    assert ws["A1"].value == "SAO KÊ TÀI KHOẢN"

    # Test CSV
    resp_csv = client.get(f"/api/v1/exports/statement.csv?account_id={a}")
    assert resp_csv.status_code == 200
    assert resp_csv.content.startswith(b"\xef\xbb\xbf")  # UTF-8 BOM
    text_content = resp_csv.content.decode("utf-8-sig")
    assert "SAO KÊ TÀI KHOẢN" in text_content
    assert "Số tiền giao dịch" in text_content


@pytest.mark.parametrize("prefix", ["=", "+", "-", "@", "\t", "\r", "\n"])
def test_statement_text_sanitizer_neutralizes_formula_and_control_prefixes(
    prefix: str,
) -> None:
    value = f"{prefix}SYNTHETIC()"
    assert _sanitize_statement_text(value) == f"'{value}"


@pytest.mark.parametrize("value", ["", "Normal text", "Tiền điện tháng Tám", "漢字"])
def test_statement_text_sanitizer_preserves_normal_and_unicode_text(value: str) -> None:
    assert _sanitize_statement_text(value) == value


def test_statement_exports_sanitize_only_untrusted_text_cells(
    client: TestClient,
) -> None:
    from openpyxl import load_workbook

    account_id = _make_account(client, "=SYNTHETIC_ACCOUNT()")
    response = client.post(
        "/api/v1/financial-events",
        json={
            "event_type": "EXPENSE",
            "transaction_date": "2026-08-02",
            "note": "@SYNTHETIC_NOTE() – tiếng Việt",
            "entries": [{"account_id": account_id, "amount": "-12.3400"}],
        },
    )
    assert response.status_code == 201

    xlsx_response = client.get(
        f"/api/v1/exports/statement.xlsx?account_id={account_id}"
    )
    assert xlsx_response.status_code == 200
    worksheet = load_workbook(io.BytesIO(xlsx_response.content)).active
    assert worksheet["B4"].value == "'=SYNTHETIC_ACCOUNT()"
    assert worksheet["D9"].value == "'@SYNTHETIC_NOTE() – tiếng Việt"
    assert worksheet["F9"].value == "-12.34"
    assert worksheet["G9"].value == "-12.34"

    csv_response = client.get(
        f"/api/v1/exports/statement.csv?account_id={account_id}"
    )
    assert csv_response.status_code == 200
    rows = list(csv.reader(io.StringIO(csv_response.content.decode("utf-8-sig"))))
    assert rows[0][1] == "'=SYNTHETIC_ACCOUNT()"
    assert rows[6][3] == "'@SYNTHETIC_NOTE() – tiếng Việt"
    assert rows[6][5] == "-12"
    assert rows[6][6] == "-12"


def test_statement_xlsx_preserves_exact_scaled_money_values(
    client: TestClient,
) -> None:
    from openpyxl import load_workbook

    account_id = _make_account(client, "Synthetic exact money")
    for event_type, amount in (
        ("INCOME", "900000000000000.1234"),
        ("INCOME", "0.0001"),
        ("EXPENSE", "-0.0001"),
        ("EXPENSE", "-12.3456"),
    ):
        response = client.post(
            "/api/v1/financial-events",
            json={
                "event_type": event_type,
                "transaction_date": "2026-08-03",
                "entries": [{"account_id": account_id, "amount": amount}],
            },
        )
        assert response.status_code == 201

    response = client.get(
        f"/api/v1/exports/statement.xlsx?account_id={account_id}"
    )
    assert response.status_code == 200
    worksheet = load_workbook(io.BytesIO(response.content)).active

    assert worksheet["B5"].value == "0 VND"
    assert worksheet["E5"].value == "899,999,999,999,987.7778 VND"
    assert worksheet["B6"].value == "+900,000,000,000,000.1235 VND"
    assert worksheet["E6"].value == "-12.3457 VND"
    assert [worksheet[f"F{row}"].value for row in range(9, 13)] == [
        "+900,000,000,000,000.1234",
        "+0.0001",
        "-0.0001",
        "-12.3456",
    ]
    assert [worksheet[f"G{row}"].value for row in range(9, 13)] == [
        "900,000,000,000,000.1234",
        "900,000,000,000,000.1235",
        "900,000,000,000,000.1234",
        "899,999,999,999,987.7778",
    ]
