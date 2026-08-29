"""Synthetic persistence coverage for the MISA export foundation."""

import datetime
import io
import os
import subprocess
from decimal import Decimal
from pathlib import Path

import pytest
from openpyxl import load_workbook
from sqlalchemy import create_engine, text
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session

from app.models import (
    Account,
    AccountEntry,
    AccountType,
    FinancialEvent,
    FinancialEventType,
    MisaExportedEvent,
    MisaExportRun,
)
from app.schemas.misa_export import (
    MisaAccountMappingCreate,
    MisaExportConfigurationCreate,
)
from app.services.misa_export import (
    MISA_BANK_STATEMENT_HEADERS,
    InvalidMisaAccountMappingError,
    MisaEventAlreadyExportedError,
    create_configuration,
    export_misa_workbook,
    list_exportable_events,
    record_export,
)


@pytest.fixture
def session(tmp_path: Path) -> Session:
    database_path = tmp_path / "synthetic.db"
    env = os.environ.copy()
    env["PF_DATABASE_PATH"] = str(database_path)
    subprocess.run(["alembic", "upgrade", "head"], check=True, env=env, capture_output=True)
    engine = create_engine(f"sqlite:///{database_path}")
    with Session(engine) as database_session:
        database_session.execute(text("PRAGMA foreign_keys=ON"))
        yield database_session
    engine.dispose()


def _account(name: str) -> Account:
    return Account(name=name, account_type=AccountType.BANK, currency="VND")


def _configuration_payload(first_id: int, second_id: int) -> MisaExportConfigurationCreate:
    return MisaExportConfigurationCreate(
        name="Synthetic MISA",
        mappings=[
            MisaAccountMappingCreate(
                source_account_id=first_id,
                target_account_code="1121",
                target_account_name="Synthetic bank",
            ),
            MisaAccountMappingCreate(
                source_account_id=second_id,
                target_account_code="1121",
                target_account_name="Synthetic bank",
            ),
        ],
    )


def test_configuration_persists_explicit_many_to_one_mapping(session: Session) -> None:
    first = _account("Synthetic A")
    second = _account("Synthetic B")
    session.add_all([first, second])
    session.commit()

    configuration = create_configuration(session, _configuration_payload(first.id, second.id))

    assert len(configuration.account_mappings) == 2
    assert {mapping.source_account_id for mapping in configuration.account_mappings} == {
        first.id,
        second.id,
    }
    assert {mapping.target_account_code for mapping in configuration.account_mappings} == {"1121"}


def test_configuration_rejects_duplicate_or_unknown_source(session: Session) -> None:
    account = _account("Synthetic")
    session.add(account)
    session.commit()
    duplicate = MisaExportConfigurationCreate(
        name="Duplicate",
        mappings=[
            MisaAccountMappingCreate(
                source_account_id=account.id, target_account_code="1", target_account_name="One"
            ),
            MisaAccountMappingCreate(
                source_account_id=account.id, target_account_code="2", target_account_name="Two"
            ),
        ],
    )
    with pytest.raises(InvalidMisaAccountMappingError):
        create_configuration(session, duplicate)

    unknown = MisaExportConfigurationCreate(
        name="Unknown",
        mappings=[
            MisaAccountMappingCreate(
                source_account_id=999, target_account_code="1", target_account_name="One"
            )
        ],
    )
    with pytest.raises(InvalidMisaAccountMappingError):
        create_configuration(session, unknown)


def test_configuration_normalizes_values_and_rejects_invalid_targets(session: Session) -> None:
    first = _account("Synthetic A")
    second = _account("Synthetic B")
    second.currency = "USD"
    session.add_all([first, second])
    session.commit()

    normalized = _configuration_payload(first.id, first.id)
    normalized.mappings = normalized.mappings[:1]
    normalized.name = "  Synthetic normalized  "
    normalized.currency = " vnd "
    normalized.mappings[0].target_account_code = " 1121 "
    normalized.mappings[0].target_account_name = " Synthetic bank "
    configuration = create_configuration(session, normalized)
    assert configuration.name == "Synthetic normalized"
    assert configuration.currency == "VND"
    assert configuration.account_mappings[0].target_account_code == "1121"
    assert configuration.account_mappings[0].target_account_name == "Synthetic bank"

    blank_target = MisaExportConfigurationCreate(
        name="Blank target",
        mappings=[
            MisaAccountMappingCreate(
                source_account_id=first.id,
                target_account_code="   ",
                target_account_name="Synthetic bank",
            )
        ],
    )
    with pytest.raises(InvalidMisaAccountMappingError, match="must not be blank"):
        create_configuration(session, blank_target)

    wrong_currency = MisaExportConfigurationCreate(
        name="Wrong currency",
        mappings=[
            MisaAccountMappingCreate(
                source_account_id=second.id,
                target_account_code="1121",
                target_account_name="Synthetic bank",
            )
        ],
    )
    with pytest.raises(InvalidMisaAccountMappingError, match="currency"):
        create_configuration(session, wrong_currency)


def test_export_history_filters_and_prevents_unintentional_reexport(session: Session) -> None:
    mapped = _account("Mapped")
    unmapped = _account("Unmapped")
    session.add_all([mapped, unmapped])
    session.flush()
    configuration = create_configuration(
        session,
        MisaExportConfigurationCreate(
            name="History",
            mappings=[
                MisaAccountMappingCreate(
                    source_account_id=mapped.id,
                    target_account_code="1121",
                    target_account_name="Synthetic bank",
                )
            ],
        ),
    )
    exportable = FinancialEvent(
        event_type=FinancialEventType.INCOME,
        transaction_date=datetime.date(2026, 8, 23),
        entries=[AccountEntry(account_id=mapped.id, amount_scaled=100_000)],
    )
    not_fully_mapped = FinancialEvent(
        event_type=FinancialEventType.TRANSFER,
        transaction_date=datetime.date(2026, 8, 23),
        entries=[
            AccountEntry(account_id=mapped.id, amount_scaled=-100_000),
            AccountEntry(account_id=unmapped.id, amount_scaled=100_000),
        ],
    )
    session.add_all([exportable, not_fully_mapped])
    session.commit()

    assert [event.id for event in list_exportable_events(session, configuration.id)] == [exportable.id]

    run = record_export(
        session,
        configuration.id,
        [exportable.id, exportable.id],
        "synthetic-misa.xlsx",
        exported_at=datetime.datetime(2026, 8, 23, tzinfo=datetime.UTC),
    )
    assert run.output_filename == "synthetic-misa.xlsx"
    assert [item.financial_event_id for item in run.exported_events] == [exportable.id]
    assert list_exportable_events(session, configuration.id) == []

    with pytest.raises(MisaEventAlreadyExportedError):
        record_export(session, configuration.id, [exportable.id], "repeat.xlsx")

    assert session.query(MisaExportRun).count() == 1
    assert session.query(MisaExportedEvent).count() == 1

    with pytest.raises(InvalidMisaAccountMappingError, match="not fully mapped"):
        record_export(session, configuration.id, [not_fully_mapped.id], "unmapped.xlsx")

    assert session.query(MisaExportRun).count() == 1
    assert session.query(MisaExportedEvent).count() == 1


def test_export_history_requires_run_and_event_configuration_to_match(session: Session) -> None:
    account = _account("Synthetic")
    session.add(account)
    session.commit()
    first_configuration = create_configuration(
        session,
        MisaExportConfigurationCreate(
            name="First history",
            mappings=[
                MisaAccountMappingCreate(
                    source_account_id=account.id,
                    target_account_code="1121",
                    target_account_name="Synthetic bank",
                )
            ],
        ),
    )
    second_configuration = create_configuration(
        session,
        MisaExportConfigurationCreate(
            name="Second history",
            mappings=[
                MisaAccountMappingCreate(
                    source_account_id=account.id,
                    target_account_code="1121",
                    target_account_name="Synthetic bank",
                )
            ],
        ),
    )
    event = FinancialEvent(
        event_type=FinancialEventType.INCOME,
        transaction_date=datetime.date(2026, 8, 23),
        entries=[AccountEntry(account_id=account.id, amount_scaled=100_000)],
    )
    session.add(event)
    session.commit()
    run = record_export(session, first_configuration.id, [event.id], "first.xlsx")

    session.add(
        MisaExportedEvent(
            export_run_id=run.id,
            configuration_id=second_configuration.id,
            financial_event_id=event.id,
        )
    )
    with pytest.raises(IntegrityError):
        session.commit()
    session.rollback()


def test_exports_structured_workbook_and_records_history(session: Session) -> None:
    first = _account("Synthetic A")
    second = _account("Synthetic B")
    session.add_all([first, second])
    session.flush()
    configuration = create_configuration(session, _configuration_payload(first.id, second.id))
    event = FinancialEvent(
        event_type=FinancialEventType.TRANSFER,
        transaction_date=datetime.date(2026, 8, 23),
        note="Synthetic transfer",
        entries=[
            AccountEntry(account_id=first.id, amount_scaled=-123_456),
            AccountEntry(account_id=second.id, amount_scaled=123_456),
        ],
    )
    session.add(event)
    session.commit()

    exported = export_misa_workbook(
        session,
        configuration.id,
        "synthetic-misa.xlsx",
        exported_at=datetime.datetime(2026, 8, 23, tzinfo=datetime.UTC),
    )

    workbook = load_workbook(io.BytesIO(exported.content), read_only=True, data_only=True)
    worksheet = workbook["Bank statement"]
    rows = list(worksheet.iter_rows(values_only=True))
    workbook.close()
    assert rows[0] == MISA_BANK_STATEMENT_HEADERS
    assert rows[1][0].date() == datetime.date(2026, 8, 23)
    assert rows[1][1].date() == datetime.date(2026, 8, 23)
    assert rows[1][2:7] == (
        f"PF-{event.id}",
        "Synthetic transfer",
        "1121",
        "Synthetic bank",
        None,
    )
    assert Decimal(str(rows[1][7])) == Decimal("12.3456")
    assert rows[1][8:] == ("VND", event.id)
    assert Decimal(str(rows[2][6])) == Decimal("12.3456")
    assert rows[2][7] is None
    assert exported.export_run.output_filename == "synthetic-misa.xlsx"
    assert list_exportable_events(session, configuration.id) == []

    with pytest.raises(ValueError, match="already exported"):
        export_misa_workbook(session, configuration.id, "repeat.xlsx", [event.id])


def test_misa_workbook_sanitizes_text_without_stringifying_dates_or_money(
    session: Session,
) -> None:
    account = _account("Synthetic")
    session.add(account)
    session.flush()
    configuration = create_configuration(
        session,
        MisaExportConfigurationCreate(
            name="Formula safety",
            mappings=[
                MisaAccountMappingCreate(
                    source_account_id=account.id,
                    target_account_code="+SYNTHETIC_CODE()",
                    target_account_name="=SYNTHETIC_NAME()",
                )
            ],
        ),
    )
    event = FinancialEvent(
        event_type=FinancialEventType.EXPENSE,
        transaction_date=datetime.date(2026, 8, 29),
        note="@SYNTHETIC_NOTE()",
        entries=[AccountEntry(account_id=account.id, amount_scaled=-123_456)],
    )
    session.add(event)
    session.commit()

    exported = export_misa_workbook(session, configuration.id, "safe.xlsx")

    workbook = load_workbook(io.BytesIO(exported.content), read_only=True, data_only=False)
    row = next(workbook["Bank statement"].iter_rows(min_row=2))
    assert row[3].value == "'@SYNTHETIC_NOTE()"
    assert row[4].value == "'+SYNTHETIC_CODE()"
    assert row[5].value == "'=SYNTHETIC_NAME()"
    assert row[0].is_date
    assert row[1].is_date
    assert row[7].data_type == "n"
    assert Decimal(str(row[7].value)) == Decimal("12.3456")
    assert row[9].data_type == "n"
    assert row[9].value == event.id
    workbook.close()
